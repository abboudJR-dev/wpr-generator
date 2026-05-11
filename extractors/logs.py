"""Logs workbook extractor.

Reads the 9 summary blocks (rows 8-15) the WPR consumes from
`DB36-EPCC-LOGS__Updated_<date>.xlsx`. Column positions vary per sheet (per
CLAUDE.md), so we hard-code the layout map.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

# Per CLAUDE.md.
LOG_SHEETS: dict[str, dict] = {
    "PQD":       {"total_col": 8, "label_col": 10, "label": "Pre-Qualification"},
    "TEC":       {"total_col": 8, "label_col": 10, "label": "Technical Submittals"},
    "NCR":       {"total_col": 6, "label_col":  8, "label": "NCRs"},
    "MIR":       {"total_col": 6, "label_col":  8, "label": "Material Inspections (MIR)"},
    "RFI":       {"total_col": 7, "label_col":  9, "label": "RFIs"},
    "WIR":       {"total_col": 7, "label_col":  9, "label": "Work Inspection Reqs (Civil)"},
    "WIR (MEP)": {"total_col": 7, "label_col":  9, "label": "Work Inspection Reqs (MEP)"},
    "DTS":       {"total_col": 7, "label_col":  9, "label": "Document Submittals"},
    "SAMPLE":    {"total_col": 7, "label_col":  9, "label": "Sample Submittals"},
}

# WPR display order
DASHBOARD_ORDER = [
    "PQD", "TEC", "MIR", "NCR", "RFI", "DTS", "WIR", "WIR (MEP)", "SAMPLE",
]


@dataclass
class CategoryStats:
    """Summary of one log category for the WPR submittal-status dashboard."""

    sheet: str
    label: str
    total: int = 0
    a: int = 0
    b: int = 0
    c: int = 0
    d: int = 0
    e: int = 0
    f: int = 0
    is_rfi: bool = False
    rfi_received: int = 0
    rfi_under_review: int = 0

    @property
    def cd(self) -> int:
        return self.c + self.d

    def display_row(self) -> list[str]:
        """Format as ['Category', 'TOTAL', 'A', 'B', 'C/D', 'E', 'F'] for the slide-7 table."""
        if self.is_rfi:
            return [
                f"{self.label} †",
                str(self.total),
                "—",
                str(self.rfi_received),
                "—",
                str(self.rfi_under_review),
                "—",
            ]
        return [
            self.label,
            str(self.total),
            str(self.a),
            str(self.b),
            str(self.cd),
            str(self.e),
            str(self.f),
        ]


@dataclass
class LogsData:
    source_path: Path
    categories: list[CategoryStats] = field(default_factory=list)

    def by_sheet(self, sheet: str) -> Optional[CategoryStats]:
        return next((c for c in self.categories if c.sheet == sheet), None)


def _to_int(v) -> int:
    if v is None:
        return 0
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0


def _read_summary_block(ws, total_col: int, label_col: int) -> dict[str, int]:
    """Walk rows 9-14 collecting {status_letter: total}, plus row 15 grand total."""
    counts: dict[str, int] = {}
    for row in range(9, 15):
        label = ws.cell(row, label_col).value
        if not label:
            continue
        letter = str(label).strip()[0].upper() if isinstance(label, str) else ""
        if letter:
            counts[letter] = _to_int(ws.cell(row, total_col).value)
    counts["TOTAL"] = _to_int(ws.cell(15, total_col).value)
    return counts


def _read_rfi_block(ws, total_col: int, label_col: int) -> dict[str, int]:
    """RFI uses R/E in rows 9-10, grand total at row 11."""
    counts: dict[str, int] = {}
    for row in range(9, 11):
        label = ws.cell(row, label_col).value
        if not label:
            continue
        letter = str(label).strip()[0].upper()
        counts[letter] = _to_int(ws.cell(row, total_col).value)
    counts["TOTAL"] = _to_int(ws.cell(11, total_col).value)
    return counts


def parse_logs(xlsx_path: str | Path) -> LogsData:
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    data = LogsData(source_path=xlsx_path)

    for sheet, cfg in LOG_SHEETS.items():
        if sheet not in wb.sheetnames:
            data.categories.append(CategoryStats(sheet=sheet, label=cfg["label"]))
            continue
        ws = wb[sheet]
        if sheet == "RFI":
            counts = _read_rfi_block(ws, cfg["total_col"], cfg["label_col"])
            data.categories.append(
                CategoryStats(
                    sheet=sheet,
                    label=cfg["label"],
                    total=counts.get("TOTAL", 0),
                    is_rfi=True,
                    rfi_received=counts.get("R", 0),
                    rfi_under_review=counts.get("E", 0),
                )
            )
        else:
            counts = _read_summary_block(ws, cfg["total_col"], cfg["label_col"])
            data.categories.append(
                CategoryStats(
                    sheet=sheet,
                    label=cfg["label"],
                    total=counts.get("TOTAL", 0),
                    a=counts.get("A", 0),
                    b=counts.get("B", 0),
                    c=counts.get("C", 0),
                    d=counts.get("D", 0),
                    e=counts.get("E", 0),
                    f=counts.get("F", 0),
                )
            )

    # Sort to dashboard display order.
    by_sheet = {c.sheet: c for c in data.categories}
    data.categories = [by_sheet[s] for s in DASHBOARD_ORDER if s in by_sheet]
    return data
